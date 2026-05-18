"""
Parse COAP 2025 IIT M.Tech GATE cutoff PDFs into a normalized long-format CSV.

Source: https://gate.iisc.ac.in/coap2025/downloads.php (per-institute reports)

Output columns:
  institute, program, paper, category, round, openScore, closeScore

Notes:
  - COAP cutoffs are GATE SCORES (not ranks).
  - Where a source only publishes a single closing cutoff, openScore == closeScore.
  - Categories are normalized to: GEN, OBC, SC, ST, EWS and *-PwD variants.

Run:  python3 scripts/parse_coap2025.py
"""
import csv
import os
import re

import openpyxl
import pandas as pd
import pdfplumber

HOME = os.path.expanduser("~")
RAW = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "gate_2025_cutoffs.csv")

NUM = re.compile(r"^(NA|N\.A\.?|CFTI|-|\d+)$", re.I)


def norm(tok):
    """A raw cell -> int score or None (NA / '-' / CFTI = no data)."""
    tok = tok.strip()
    if (tok in ("-", "") or tok.upper().startswith("NA")
            or tok.upper().startswith("N.A") or tok.upper() == "CFTI"):
        return None
    try:
        return int(tok)
    except ValueError:
        return None


def split_trailing(line, n):
    """Split a line into (label, [last n score-ish tokens])."""
    parts = line.split()
    if len(parts) < n:
        return None, None
    cells = parts[-n:]
    if not all(NUM.match(c) for c in cells):
        return None, None
    label = " ".join(parts[: len(parts) - n]).strip()
    return label, cells


def parse_iith(rows):
    """IIT Hyderabad: round-wise, 8 categories x (Low High) = 16 trailing cells."""
    cats = ["GEN", "GEN-PwD", "OBC", "OBC-PwD", "EWS", "EWS-PwD", "SC", "ST"]
    with pdfplumber.open(os.path.join(RAW, "iith.pdf")) as pdf:
        rnd = 1
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                m = re.match(r"^Round\s+(\d+)", line.strip())
                if m:
                    rnd = int(m.group(1))
                    continue
                if line.strip().startswith(("GN ", "Program", "Low High")):
                    continue
                label, cells = split_trailing(line, 16)
                if not label:
                    continue
                for i, cat in enumerate(cats):
                    lo, hi = norm(cells[2 * i]), norm(cells[2 * i + 1])
                    if lo is None and hi is None:
                        continue
                    rows.append({
                        "institute": "IIT Hyderabad", "program": label, "paper": "",
                        "category": cat, "round": rnd,
                        "openScore": lo if lo is not None else hi,
                        "closeScore": hi if hi is not None else lo,
                    })


def parse_jodhpur(rows):
    """IIT Jodhpur: round-wise, 6 categories x (Low High) = 12 trailing cells.
    Program is the leading code token (BBE, CSE, ...)."""
    cats = ["GEN", "OBC", "SC", "ST", "EWS", "PwD"]
    with pdfplumber.open(os.path.join(RAW, "iitjodhpur.pdf")) as pdf:
        rnd = 1
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                m = re.match(r"^Round\s+(\d+)", line.strip())
                if m:
                    rnd = int(m.group(1))
                    continue
                label, cells = split_trailing(line, 12)
                if not label or " " in label:  # program code is a single token
                    continue
                for i, cat in enumerate(cats):
                    lo, hi = norm(cells[2 * i]), norm(cells[2 * i + 1])
                    if lo is None and hi is None:
                        continue
                    rows.append({
                        "institute": "IIT Jodhpur", "program": label, "paper": "",
                        "category": cat, "round": rnd,
                        "openScore": lo if lo is not None else hi,
                        "closeScore": hi if hi is not None else lo,
                    })


def parse_guwahati(rows):
    """IIT Guwahati: '<program> <Mcode> <CATEGORY> <score>' one row per pair."""
    cmap = {"OPEN": "GEN", "OBC-NCL": "OBC", "GEN-EWS": "EWS", "SC": "SC", "ST": "ST"}
    pat = re.compile(r"^(.+?)\s+(M\d{4})\s+([A-Z-]+)\s+(\d+)$")
    with pdfplumber.open(os.path.join(RAW, "iitg.pdf")) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                m = pat.match(line.strip())
                if not m:
                    continue
                cat = cmap.get(m.group(3))
                if not cat:
                    continue
                rows.append({
                    "institute": "IIT Guwahati", "program": m.group(1).strip(),
                    "paper": m.group(2), "category": cat, "round": 0,
                    "openScore": int(m.group(4)), "closeScore": int(m.group(4)),
                })


def parse_gandhinagar(rows):
    """IIT Gandhinagar: 10 interleaved cells = (cat, cat-PwD) x 5."""
    cats = ["GEN", "GEN-PwD", "EWS", "EWS-PwD", "OBC",
            "OBC-PwD", "SC", "SC-PwD", "ST", "ST-PwD"]
    with pdfplumber.open(os.path.join(RAW, "iitgn.pdf")) as pdf:
        for line in (pdf.pages[0].extract_text() or "").split("\n"):
            label, cells = split_trailing(line, 10)
            if not label or not label[0:1].isalpha() or label.startswith(
                    ("MTech", "Category", "GEN", "Note", "INDIAN")):
                continue
            for i, cat in enumerate(cats):
                s = norm(cells[i])
                if s is None:
                    continue
                rows.append({
                    "institute": "IIT Gandhinagar", "program": label,
                    "paper": "", "category": cat, "round": 0,
                    "openScore": s, "closeScore": s,
                })


def _cluster(vals, tol=15):
    """Cluster sorted numeric values; return cluster means."""
    out, cur = [], [vals[0]]
    for v in vals[1:]:
        if v - cur[-1] <= tol:
            cur.append(v)
        else:
            out.append(sum(cur) / len(cur))
            cur = [v]
    out.append(sum(cur) / len(cur))
    return out


def parse_ropar(rows):
    """IIT Ropar: one PDF per round. Layout varies between rounds, so words
    are assigned to category columns by x-position. 12 column anchors come
    from the 'Opening'/'Closing' header words; pairs map to
    [GEN OBC SC ST EWS PwD]."""
    cats = ["GEN", "OBC", "SC", "ST", "EWS", "PwD"]
    for rnd in range(1, 11):
        path = os.path.join(RAW, f"ropar_r{rnd}.pdf")
        if not os.path.exists(path):
            continue
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                w = page.extract_words()
                oc = sorted((x["x0"] + x["x1"]) / 2 for x in w
                            if x["text"] in ("Opening", "Closing"))
                if len(oc) < 12:
                    continue
                cols = _cluster(oc)
                if len(cols) != 12:
                    continue
                first = min(cols)
                anchors = sorted((x for x in w
                                  if x["text"].upper().startswith("M.TECH")
                                  and x["x1"] < first - 35),
                                 key=lambda x: x["top"])
                for k, an in enumerate(anchors):
                    y0 = an["top"] - 4
                    y1 = anchors[k + 1]["top"] - 4 if k + 1 < len(anchors) else 1e9
                    blk = [x for x in w if y0 <= x["top"] < y1]
                    name = " ".join(
                        x["text"] for x in sorted(blk, key=lambda x: (x["top"], x["x0"]))
                        if x["x1"] < first - 35 and not x["text"].isdigit()
                        and x["text"] not in ("Opening", "Closing", "GATE",
                                               "SCORE", "PROGRAME", "NAME", "SR", "NO"))
                    name = " ".join(name.split())
                    slots = {}
                    for x in blk:
                        if not x["text"].lstrip("-").isdigit():
                            continue
                        cx = (x["x0"] + x["x1"]) / 2
                        if cx < first - 35:
                            continue
                        ci = min(range(12), key=lambda i: abs(cols[i] - cx))
                        if abs(cols[ci] - cx) <= 28 and ci not in slots:
                            slots[ci] = int(x["text"])
                    if not name or not slots:
                        continue
                    for i, cat in enumerate(cats):
                        lo, hi = slots.get(2 * i), slots.get(2 * i + 1)
                        if lo is None and hi is None:
                            continue
                        rows.append({
                            "institute": "IIT Ropar", "program": name,
                            "paper": "", "category": cat, "round": rnd,
                            "openScore": lo if lo is not None else hi,
                            "closeScore": hi if hi is not None else lo,
                        })


def parse_dharwad(rows):
    """IIT Dharwad: per-program blocks; 'Round N'/'Additional Round' rows
    with 12 cells = (Opening Closing) x [GEN EWS OBC SC ST PwD]."""
    cats = ["GEN", "EWS", "OBC", "SC", "ST", "PwD"]
    with pdfplumber.open(os.path.join(RAW, "iitdh.pdf")) as pdf:
        prog = ""
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                s = line.strip()
                m = re.match(r"^M\.Tech\._?\s*(\S+)", s)
                if m:
                    prog = m.group(1)
                    continue
                rm = re.match(r"^(Round\s+\d+|Additional Round)\b", s)
                if not rm or not prog:
                    continue
                rnd = 11 if s.startswith("Additional") else int(
                    re.search(r"\d+", rm.group(1)).group())
                _, cells = split_trailing(s, 12)
                if not cells:
                    continue
                for i, cat in enumerate(cats):
                    lo, hi = norm(cells[2 * i]), norm(cells[2 * i + 1])
                    if lo is None and hi is None:
                        continue
                    rows.append({
                        "institute": "IIT Dharwad", "program": prog,
                        "paper": "", "category": cat, "round": rnd,
                        "openScore": lo if lo is not None else hi,
                        "closeScore": hi if hi is not None else lo,
                    })


def parse_indore(rows):
    """IIT Indore: round-wise; 14 cells = (Max Min) x
    [General Gen-PwD EWS OBC OBC-PwD SC ST]. Max=opening, Min=closing."""
    cats = ["GEN", "GEN-PwD", "EWS", "OBC", "OBC-PwD", "SC", "ST"]
    with pdfplumber.open(os.path.join(RAW, "iiti.pdf")) as pdf:
        rnd, pending = 0, ""
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                s = line.strip()
                m = re.search(r"Round[-\s]+(\d+)", s)
                if m and "COAP" in s:
                    rnd = int(m.group(1))
                    continue
                label, cells = split_trailing(s, 14)
                if not label:
                    if s and not s.startswith(("General", "Programs", "Max of",
                                               "GATE", "Score", "INDIAN")):
                        pending = (pending + " " + s).strip()
                    continue
                program = (pending + " " + label).strip()
                pending = ""
                if program.startswith(("General", "Programs", "Max", "GATE", "Score")):
                    continue
                for i, cat in enumerate(cats):
                    hi, lo = norm(cells[2 * i]), norm(cells[2 * i + 1])
                    if lo is None and hi is None:
                        continue
                    rows.append({
                        "institute": "IIT Indore", "program": program,
                        "paper": "", "category": cat, "round": rnd,
                        "openScore": hi if hi is not None else lo,
                        "closeScore": lo if lo is not None else hi,
                    })


def parse_patna(rows):
    """IIT Patna (Collegedunia, third-party aggregator — not official COAP):
    5 course tables (17 courses x Round 1-10 closing marks), in category
    order GEN, OBC, SC, ST, EWS."""
    path = os.path.join(RAW, "patna.html")
    if not os.path.exists(path):
        return
    cat_order = ["GEN", "OBC", "SC", "ST", "EWS"]
    tbls = [t for t in pd.read_html(path)
            if t.shape[0] >= 10 and str(t.columns[0]).startswith("Courses")
            and "Round 1 (Closing marks)" in list(t.columns)]
    for ci, tbl in enumerate(tbls[:5]):
        cat = cat_order[ci]
        for _, r in tbl.iterrows():
            program = str(r.iloc[0]).strip()
            if not program or program.lower() == "nan":
                continue
            for rnd in range(1, 11):
                col = f"Round {rnd} (Closing marks)"
                if col not in tbl.columns:
                    continue
                v = norm(str(r[col]).strip())
                if v is None:
                    continue
                rows.append({
                    "institute": "IIT Patna", "program": program,
                    "paper": "", "category": cat, "round": rnd,
                    "openScore": v, "closeScore": v,
                })


def parse_goa(rows):
    """IIT Goa: round-wise; '<code>' + 12 cells = (Min Max) x
    [UR OBC SC ST EWS PwD]."""
    cats = ["GEN", "OBC", "SC", "ST", "EWS", "PwD"]
    with pdfplumber.open(os.path.join(RAW, "iitgoa.pdf")) as pdf:
        rnd = 0
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                s = line.strip()
                m = re.match(r"^Round-(\d+)", s)
                if m:
                    rnd = int(m.group(1))
                    continue
                label, cells = split_trailing(s, 12)
                if not label or " " in label:
                    continue
                for i, cat in enumerate(cats):
                    lo, hi = norm(cells[2 * i]), norm(cells[2 * i + 1])
                    if lo is None and hi is None:
                        continue
                    rows.append({
                        "institute": "IIT Goa", "program": label, "paper": "",
                        "category": cat, "round": rnd,
                        "openScore": lo if lo is not None else hi,
                        "closeScore": hi if hi is not None else lo,
                    })


def parse_delhi(rows):
    """IIT Delhi (GATE Score.xlsm): sheets 'Main Round' / 'Spot Round';
    cols Code, Discipline, GEN, EWS, OBC, SC, ST, PH(PwD).
    Non-numeric cells ('No Selected', 'NO Selection', ...) = no data."""
    path = os.path.join(HOME, "Downloads", "GATE Score.xlsm")
    if not os.path.exists(path):
        return
    cats = ["GEN", "EWS", "OBC", "SC", "ST", "PwD"]
    rnd_of = {"Main Round": 1, "Spot Round": 2}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        rnd = rnd_of.get(sn)
        if rnd is None:
            continue
        for row in wb[sn].iter_rows(values_only=True):
            if not row or not row[0] or str(row[0]).strip() == "Code":
                continue
            code = str(row[0]).strip()
            paper = str(row[1]).strip() if row[1] else ""
            for i, cat in enumerate(cats):
                v = row[2 + i] if 2 + i < len(row) else None
                if not isinstance(v, (int, float)) or v <= 0:
                    continue
                rows.append({
                    "institute": "IIT Delhi", "program": code,
                    "paper": paper, "category": cat, "round": rnd,
                    "openScore": int(v), "closeScore": int(v),
                })
    wb.close()


def parse_palakkad(rows):
    """IIT Palakkad: table; category cells hold one or more '<PAPER>-<score>'
    entries. The lowest qualifying score per cell is recorded."""
    cats = ["GEN", "EWS", "OBC", "SC", "ST"]
    with pdfplumber.open(os.path.join(RAW, "iitpkd.pdf")) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables():
                for row in tbl:
                    if not row or not row[0] or str(row[0]).startswith("M. Tech"):
                        continue
                    program = " ".join(str(row[0]).split())
                    for i, cat in enumerate(cats):
                        cell = row[2 + i]
                        if not cell:
                            continue
                        nums = [int(x) for x in re.findall(r"[A-Z]+-(\d+)", str(cell))]
                        if not nums:
                            continue
                        rows.append({
                            "institute": "IIT Palakkad", "program": program,
                            "paper": "", "category": cat, "round": 0,
                            "openScore": min(nums), "closeScore": min(nums),
                        })


def parse_iisc(rows):
    """IISc Bangalore: round-wise; per department, paper rows with
    16 cells = (Low High) x [UR OBC SC ST EWS WQ PwD KM].
    UR/OBC/SC/ST/EWS/PwD are kept; WQ and KM are not standard categories."""
    idx = {"GEN": 0, "OBC": 2, "SC": 4, "ST": 6, "EWS": 8, "PwD": 12}
    head = re.compile(r"^[A-Z]{2,4}(\s+(-|\d+)){16}$")
    with pdfplumber.open(os.path.join(RAW, "iisc_gate.pdf")) as pdf:
        rnd, dept = 0, ""
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                s = line.strip()
                m = re.match(r"^ROUND\s+(\d+)", s)
                if m:
                    rnd = int(m.group(1))
                    continue
                if not s or s.startswith(("M.Tech", "Operated", "UR ",
                                          "DEPARTMENT", "DISCIPLINE")):
                    continue
                if not head.match(s):
                    dept = s
                    continue
                parts = s.split()
                paper, cells = parts[0], parts[1:]
                for cat, j in idx.items():
                    lo, hi = norm(cells[j]), norm(cells[j + 1])
                    if lo is None and hi is None:
                        continue
                    rows.append({
                        "institute": "IISc Bangalore",
                        "program": dept, "paper": paper, "category": cat,
                        "round": rnd,
                        "openScore": lo if lo is not None else hi,
                        "closeScore": hi if hi is not None else lo,
                    })


def parse_kanpur(rows):
    """IIT Kanpur: one xlsx per round; sheets = <cat>_<NonDAP|DAP>;
    score sits under its GATE-paper column. Lowest qualifying score per
    program/category/round is recorded (with the paper that produced it)."""
    smap = {"GN": "GEN", "EWS": "EWS", "OBC_NCL": "OBC", "SC": "SC", "ST": "ST"}
    for rnd in range(1, 11):
        path = os.path.join(RAW, f"iitk_r{rnd}.xlsx")
        if not os.path.exists(path):
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            base, _, kind = sheet.rpartition("_")
            cat = smap.get(base)
            if not cat:
                continue
            if kind == "DAP":
                cat += "-PwD"
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            papers = list(next(it))  # header row
            for row in it:
                if not row or not row[0]:
                    continue
                best, best_paper = None, ""
                for ci in range(2, len(row)):
                    v = row[ci]
                    if isinstance(v, (int, float)) and v > 0:
                        if best is None or v < best:
                            best, best_paper = int(v), str(papers[ci])
                if best is None:
                    continue
                rows.append({
                    "institute": "IIT Kanpur", "program": str(row[0]).strip(),
                    "paper": best_paper, "category": cat, "round": rnd,
                    "openScore": best, "closeScore": best,
                })
        wb.close()


def parse_jammu(rows):
    """IIT Jammu: single-page table; columns Gen OBC-NCL EWS SC ST PwD.
    Names and their scores sit on slightly different baselines, so numbers
    are bucketed to programs by y-band and to categories by header x."""
    catmap = {"Gen": "GEN", "OBC-NCL": "OBC", "EWS": "EWS",
              "SC": "SC", "ST": "ST", "PwD": "PwD"}
    with pdfplumber.open(os.path.join(RAW, "iitjammu.pdf")) as pdf:
        p = pdf.pages[0]
        w = p.extract_words()
        hdr = {x["text"]: (x["x0"] + x["x1"]) / 2 for x in w
               if x["text"] in catmap}
        # Cluster left-column words into program names by line (top).
        left = sorted((x for x in w if x["x0"] < 175
                       and not x["text"].isdigit()
                       and x["text"] != "Programme"),
                      key=lambda x: x["top"])
        names, cur, ctop = [], [], None
        for x in left:
            if ctop is None or abs(x["top"] - ctop) <= 8:
                cur.append(x)
                ctop = x["top"] if ctop is None else ctop
            else:
                names.append(cur)
                cur, ctop = [x], x["top"]
        if cur:
            names.append(cur)
        anchors = [(min(g, key=lambda x: x["x0"])["top"],
                    " ".join(t["text"] for t in sorted(g, key=lambda x: x["x0"])))
                   for g in names]
        anchors.sort()
        for i, (top, name) in enumerate(anchors):
            y0 = top - 14
            y1 = anchors[i + 1][0] - 14 if i + 1 < len(anchors) else 1e9
            for x in w:
                if not x["text"].isdigit() or not (y0 <= x["top"] < y1):
                    continue
                cx = (x["x0"] + x["x1"]) / 2
                cat = min(hdr, key=lambda c: abs(hdr[c] - cx))
                if abs(hdr[cat] - cx) > 30:
                    continue
                v = int(x["text"])
                rows.append({
                    "institute": "IIT Jammu", "program": name,
                    "paper": "", "category": catmap[cat], "round": 0,
                    "openScore": v, "closeScore": v,
                })


def parse_ism(rows):
    """IIT ISM Dhanbad: '<program> <ID> <UR OBC EWS SC ST> [PwD...]'.
    Trailing PwD columns are sparsely populated with ambiguous alignment,
    so only the reliable first 5 (non-PwD) categories are taken."""
    cats = ["GEN", "OBC", "EWS", "SC", "ST"]
    with pdfplumber.open(os.path.join(RAW, "iitism.pdf")) as pdf:
        pending = ""
        for line in (pdf.pages[0].extract_text() or "").split("\n"):
            s = line.strip()
            if not s or s.startswith(("PG-", "INDIAN", "GATE CUT", "Program ")):
                continue
            toks = s.split()
            n = 0
            while n < len(toks) and toks[-1 - n].isdigit():
                n += 1
            if n < 5 or n + 1 >= len(toks):
                pending = (pending + " " + s).strip()
                continue
            nums = toks[len(toks) - n:]
            program = (pending + " " + " ".join(toks[: len(toks) - n - 1])).strip()
            pending = ""
            for i, cat in enumerate(cats):
                v = norm(nums[i])
                if v is None:
                    continue
                rows.append({
                    "institute": "IIT ISM Dhanbad", "program": program,
                    "paper": toks[len(toks) - n - 1], "category": cat, "round": 10,
                    "openScore": v, "closeScore": v,
                })


def parse_bhilai(rows):
    """IIT Bhilai: 10 cells; program names wrap across lines."""
    cats = ["EWS", "EWS-PwD", "OBC", "OBC-PwD", "SC",
            "SC-PwD", "ST", "ST-PwD", "GEN", "GEN-PwD"]
    with pdfplumber.open(os.path.join(RAW, "iitbhilai.pdf")) as pdf:
        pending = ""
        for line in (pdf.pages[0].extract_text() or "").split("\n"):
            if line.strip().startswith(("Category", "Discipline", "Note", "EWS",
                                        "OBC", "PWD", "PwD", "(NCL)")) \
                    or re.match(r"^\d+\.", line.strip()):
                continue
            label, cells = split_trailing(line, 10)
            if not label:
                pending = (pending + " " + line.strip()).strip()
                continue
            program = (pending + " " + label).strip()
            pending = ""
            for i, cat in enumerate(cats):
                s = norm(cells[i])
                if s is None:
                    continue
                rows.append({
                    "institute": "IIT Bhilai", "program": program,
                    "paper": "", "category": cat, "round": 0,
                    "openScore": s, "closeScore": s,
                })


def parse_madras(rows):
    """IIT Madras: table; col0 = program code, then 10 ordered score tokens
    (GEN EWS OBC SC ST + same PwD). Cell positions drift, so filter in order."""
    cats = ["GEN", "EWS", "OBC", "SC", "ST",
            "GEN-PwD", "EWS-PwD", "OBC-PwD", "SC-PwD", "ST-PwD"]
    code = re.compile(r"^[A-Z]{2}\d[A-Z]?\d?\*?$")
    with pdfplumber.open(os.path.join(RAW, "iitm.pdf")) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables():
                for row in tbl:
                    if not row or not row[0] or not code.match(str(row[0]).strip()):
                        continue
                    toks = [str(c).strip() for c in row[3:]
                            if c is not None and str(c).strip() != ""]
                    toks = [t for t in toks if t == "-" or t.isdigit()]
                    if len(toks) < 5:
                        continue
                    for i, cat in enumerate(cats[:len(toks)]):
                        s = norm(toks[i])
                        if s is None:
                            continue
                        rows.append({
                            "institute": "IIT Madras", "program": row[0].strip(),
                            "paper": "", "category": cat, "round": 0,
                            "openScore": s, "closeScore": s,
                        })


def parse_single(rows, pdf_name, institute, cats, ncols, label_from):
    """Generic single-score-per-category parser.

    cats: ordered category names matching the ncols trailing numeric cells.
    label_from: callable(parts_before_cells) -> program string.
    """
    with pdfplumber.open(os.path.join(RAW, pdf_name)) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").split("\n"):
                label, cells = split_trailing(line, ncols)
                if not label:
                    continue
                program = label_from(label)
                if not program:
                    continue
                for i, cat in enumerate(cats):
                    s = norm(cells[i])
                    if s is None:
                        continue
                    rows.append({
                        "institute": institute, "program": program, "paper": "",
                        "category": cat, "round": 0,
                        "openScore": s, "closeScore": s,
                    })


def main():
    rows = []
    parse_iith(rows)
    parse_jodhpur(rows)
    parse_guwahati(rows)
    parse_gandhinagar(rows)
    parse_madras(rows)
    parse_bhilai(rows)
    parse_ism(rows)
    parse_jammu(rows)
    parse_kanpur(rows)
    parse_palakkad(rows)
    parse_iisc(rows)
    parse_goa(rows)
    parse_delhi(rows)
    parse_patna(rows)
    parse_dharwad(rows)
    parse_indore(rows)
    parse_ropar(rows)

    # IIT KGP: <ProgramCode> + Major(GEN EWS OBC SC ST) + Minor(GEN EWS OBC SC ST) + PWD
    parse_single(
        rows, "iitkgp.pdf", "IIT Kharagpur",
        ["GEN", "EWS", "OBC", "SC", "ST"], 11,
        lambda lbl: lbl if lbl and not lbl.startswith(("Program", "JMP")) else None,
    )

    # IIT BHU: ... <PaperCode> + Non-PwD(GEN OBC-NCL SC ST EWS) + PwD(GEN OBC-NCL SC ST EWS)
    parse_single(
        rows, "iitbhu.pdf", "IIT BHU",
        ["GEN", "OBC", "SC", "ST", "EWS"], 10,
        lambda lbl: lbl if lbl and lbl[0:1].isalpha()
        and not lbl.startswith(("Program", "Discipline", "M.Tech", "1)", "2)")) else None,
    )

    # Drop rows misaligned by program names that wrap across PDF lines.
    rows = [r for r in rows if r["closeScore"] and 1 <= r["closeScore"] <= 1000]

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "institute", "program", "paper", "category",
            "round", "openScore", "closeScore"])
        w.writeheader()
        w.writerows(rows)

    by_inst = {}
    for r in rows:
        by_inst[r["institute"]] = by_inst.get(r["institute"], 0) + 1
    print(f"Wrote {len(rows)} rows -> {os.path.relpath(OUT)}")
    for k, v in sorted(by_inst.items()):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
